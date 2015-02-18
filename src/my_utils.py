# coding: utf-8

# my_utils for Intro to Data Science with Python
# Author: Kat Chuang
# Created: Nov 2014

# --------------------------------------
## Stage 2 begin

import csv 

# open a file and return a double list
def open_with_csv(filename, d='\t'):
    uuids = []
    with open(filename, encoding='utf-8') as tsvin:
        tsvin = csv.reader(tsvin, delimiter=d)
        for row in tsvin:
            uuids.append(row)
    return uuids

#2.a size
def number_of_records(data_sample):
    return len(data_sample)


#2.b calculate sum
def calculate_sum(data_sample):
	total = 0
	for row in data_sample[1:]:
	    price = float(row[2])
	    total += price
	return total

#2.c 1 # Find the average price
def find_average(data_sample, headers=False):
    total = calculate_sum(data_sample)
    size = number_of_records(data_sample)
    if headers:
        total -= 1

    average = total / size
    
    # return average   
    
    return '{:03.2f}'.format(average) #nicely formatted to 2 decimals

#2.d. Max, Min
# Find the maximum price
def find_max(theData, col):
    tempList = []
    
    for row in theData:
        price = float(row[col])
        tempList.append(price)
    return max(tempList)

# Find the minimum price
def find_min(theData, col):
    tempList = []

    for row in theData:
        price = float(row[col])
        tempList.append(price)
    return min(tempList)

# Find the max or min price
def find_max_min(the_data, col, m):
    tempList = [], val = 0

    for row in the_data:
        price = float(row[col])
        tempList.append(price)

        if m == "max": 
            val = max(tempList)
        elif m == "min":
            val = min(tempList)
        else: # hopefully we don’t come to this
            pass 

    return val


## Stage 2 end

# --------------------------------------

## Stage 3 begin
# Stage 3: Cleaning data

def create_bool_field_from_search_term(data_sample, search_term):
    new_array = []
    new_array.append(data_sample[0].append("cashmere"))

    for row in data_sample[1:]:
        new_bool_field=False
        if search_term in row[7]:
            new_bool_field=True
        
        row.append(new_bool_field)
        new_array.append(row)

    return new_array 

#filter by boolean
def filter_col_by_bool(the_data, col):
    filtered_rows = []
    
    for row in the_data[1:]:
        if row[col]:
            filtered_rows.append(row)
            
    return filtered_rows

# Filter rows were columns match a string data type
def filter_col_by_string(the_data, field, filter_condition):
    filtered_rows = []
    
    #find index of field in first row
    col = int(the_data[0].index(field))
    filtered_rows.append(the_data[0])

    for row in the_data[1:]:
        if row[col] == filter_condition:
            filtered_rows.append([str(x).encode('utf8') for x in row])
            
    return filtered_rows

# Filter rows were columns match a float data type
def filter_col_by_float(the_data, field, direction, filter_condition):
    filtered_rows = []
    
    #find index of field in first row
    col = int(the_data[0].index(field))
    cond = float(filter_condition)
    
    for row in the_data[1:]:
        element = float(row[col])
        
        if direction == "<":
            if element < cond: filtered_rows.append(row)
                
        elif direction == "<=":
            if element <= cond: filtered_rows.append(row)

        elif direction == ">":
            if element > cond: filtered_rows.append(row)

        elif  direction == ">=":
            if element >= cond: filtered_rows.append(row)
                
        elif  direction == "==":
            if element == cond: filtered_rows.append(row)
        else:
            pass
        
    return filtered_rows


## Stage 3 end

# --------------------------------------

## Stage 4 begin

#4.a csv
def write_to_file(filename, data_sample):
    example = csv.writer(open(filename, 'w', newline='', encoding='utf-8'))
    example.writerows(data_sample)

#4.b more functions
def write_brand_and_price_to_file(filename, data_sample):

    # confirm that the columns only have two columns, otherwise take the two fields
    num_fields = len(data_sample[0])
    
    brand_field_index = 5 #int(dataSample[0].index("brand"))
    price_field_index = 2 #int(dataSample[0].index("priceLabel"))
    
    #if numFields > 2:
    new_array = []
    for record in data_sample:
        new_record = [None] * 2
        new_record[0] = record[brand_field_index]
        new_record[1] = record[price_field_index]
        new_array.append(new_record)

    # write the file
    write_to_file(filename, new_array)  

def write_min_max_csv(filename, data_sample):

    #find min & max price from data_sample
    min = find_max_min(data_sample, 2, "min")
    max = find_max_min(data_sample, 2, "max")

    new_array = []
    for record in data_sample:

        if (float(record[2]) == min) or (float(record[2]) == max): 
            new_array.append(record)

    write_to_file(filename, new_array)  


#csv with just 2 columns
def write_two_cols(filename, data_sample, col1, col2):    

    #if numFields > 2:
    new_array = []
    for record in data_sample:
        new_record = [None] * 2
        new_record[0] = record[col1]
        new_record[1] = record[col2]
        new_array.append(new_record)

    # write the file
    write_to_file(filename, new_array)

#csv with sorted prices
# ascending is going up i.e. A-Z 
# descending is going down i.e. Z-A 

def write_sorted_prices(filename, data_sample, order="ascending"):

    if order == "descending":
        data_sample.sort(key=lambda x: float(x[2]), reverse=False)
    else:
        data_sample.sort(key=lambda x: float(x[2]), reverse=True)

    write_to_file(filename, data_sample) 

#append another file
def write_append_file(filename, new_data_to_add):
    with open(filename, "a", encoding='utf-8') as myfile:
        for row in new_data_to_add:
            myfile.write(str(row))


#4.c export to excel
from openpyxl import Workbook
from openpyxl.cell import get_column_letter

def save_spreadsheet(filename, data_sample):
    
    wb = Workbook()
    ws = wb.active

    rowIndex = 1
    for rows in data_sample:
        colIndex = 1 
        for field in rows:
            colIndex2 = get_column_letter(colIndex)
            ws.cell('%s%s'%(colIndex2, rowIndex)).value = field
            colIndex +=1
        rowIndex += 1
    
    wb.save(filename)

## Stage 4 end

# --------------------------------------

## Stage 5 begin
import matplotlib.pyplot as plt

#5.a Line charts
def create_line_chart(sample, title, exported_figure_filename):
    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)

    prices = sorted(map(int, sample))
    x_axis_ticks = list( range(len(sample)) )
    ax.plot(x_axis_ticks, prices, label='price points', linewidth=2)
    ax.set_title(title)
    ax.set_xlabel('Tie Price ($)')
    ax.set_ylabel('Number of Ties')
    ax.set_xlim([0,len(sample)])
    fig.savefig(exported_figure_filename)


#5.b bar charts
def create_bar_chart(price_groups, exported_figure_filename):
    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)
    colors=plt.rcParams['axes.color_cycle']
     
    for group in price_groups:
        ax.bar(group, price_groups[group], color=colors[group%len(price_groups)])

    labels = ["$0-50", "$50-100", "$100-150", "$150-200", "$200-250", "$250+"]
    ax.legend(labels)

    ax.set_title('Amount of Ties at price points')
    ax.set_xlabel('Tie Price ($)')
    ax.set_xticklabels(labels, ha='left')
    ax.set_xticks( range(1, len(price_groups)+1) )
    ax.set_ylabel('Number of Ties')

    plt.grid(True)
    fig.savefig(exported_figure_filename)


#5.c tables

import prettytable
from prettytable import PrettyTable

def my_table(data_sample):

    x = PrettyTable(data_sample[0])
    x.align["City name"] = "l" 
    x.padding_width = 1 

    for row in data_sample[1:]:
        x.add_row(row)
    
    print(x)


def create_table(data_sample, price_groups, brand_names, columns, exported_figure_filename):
    tup = build_table_text(data_sample, brand_names)

    fig = plt.figure()
    ax = fig.add_subplot(1, 1, 1)

    for group in price_groups:
        plt.bar(group, price_groups[group]) #color=colors[group%len(price_groups)]

    ax.table(cellText=tup[0], colLabels=columns, rowLabels=tup[1], loc='bottom')
    ax.text(-1.3, 0, 'Discounted Ties Brands', size=12, horizontalalignment='left', verticalalignment='top')
    ax.tick_params(
        axis='x',          # changes apply to the x-axis
        which='both',      # both major and minor ticks are affected
        labelbottom='off') # labels along the bottom edge are off

    fig.savefig(exported_figure_filename, dpi=400, bbox_inches='tight')


from collections import Counter
def group_prices_by_range(prices_in_float):
    
    tally = Counter()

    for item in prices_in_float:
        bucket = 0
        rounded_price = round(item, -1)
        if rounded_price >= 0 and rounded_price <= 50:
            bucket = 1
        elif rounded_price >= 50 and rounded_price <= 100:
            bucket = 2
        elif rounded_price >= 100 and rounded_price <= 150:
            bucket = 3
        elif rounded_price >= 150 and rounded_price <= 200:
            bucket = 4
        elif rounded_price >= 200 and rounded_price <= 250:
            bucket = 5
        elif rounded_price >= 250:
            bucket = 6
        else:
            bucket = 7

        tally[bucket] += 1
    return tally

def count_prices_for_brands(data_sample, brand, min_price, max_price):
    count = 0
    for row in data_sample: 
        if str(row[0]) == str(brand):
            if float(row[1]) < max_price:
                if float(row[1]) > min_price:
                    count += 1
    return count

def build_table_text(data_sample, brands):  
    cell_text = []
    row_text = []

    unique_brand_list = sorted(set(brands))
    for b in unique_brand_list:
        b = bytes.decode(b)
        temp_row = [] 
        group1 = count_prices_for_brands(data_sample, b, 0, 50.00)
        group2 = count_prices_for_brands(data_sample, b, 50.00, 100.00)
        group3 = count_prices_for_brands(data_sample, b, 100.00, 150.00)
        group4 = count_prices_for_brands(data_sample, b, 150.00, 200.00)
        group5 = count_prices_for_brands(data_sample, b, 200.00, 250.00)
        group6 = count_prices_for_brands(data_sample, b, 250.00, 1000.00)
        row_list = [group1, group2, group3, group4, group5, group6]
        temp_row.extend(row_list) 
        
        if group1 > 0:
            if any(x >= group1 for x in row_list[1:]):
                cell_text.append(temp_row)
                row_text.append(b)

    return (cell_text, row_text)

def print_brand_avg_min(name, data_from_csv):
    tie_sample = filter_col_by_string(data_from_csv, "brandName", name)
    avg_price = calculate_sum(tie_sample) / len(tie_sample)
    min_price = find_min(tie_sample[1:], 2)
    print("{2} Average: ${0:6.2f}; Min: ${1:.2f}".format(avg_price, min_price, name))


## Stage 5 end

# --------------------------------------

## Stage 6 begin

import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

def plot_minimal_graph(tally, columns, *args):
    plt.style.use('bmh')
    fig = plt.figure(dpi=200) 
    colors=plt.rcParams['axes.color_cycle']
        
    # --- White background to use less printer ink --- #
    ax = plt.subplot(111,axisbg='white')
        
    # Plot bars and create text labels for the table
    for priceBucket in tally:
        ax.bar(priceBucket, tally[priceBucket], color=colors[priceBucket%len(tally)])
        ax.annotate(r"%d" % (tally[priceBucket]),
                   (priceBucket+0.2, tally[priceBucket]), 
                   va="bottom", ha="center")


    # --- Include a legend  --- #
    ax.legend(columns)
      
    # --- Remove distracting lines on top, left, and right --- #
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_visible(False)

    # --- Remove distracting tick marks  --- #
    ax.yaxis.set_ticks_position('none')
    ax.xaxis.set_ticks_position('none')


    # --- Add chart title and axes labels --- #          
    plt.xlabel('Tie Price', fontsize = 13)
    plt.ylabel('Number of Ties', fontsize = 13)
    plt.title('Chart #1')


    # --- Add labels to bars along x axes --- #
    x = range(1, len(tally)+1)
    plt.xticks(x, columns, rotation='horizontal', ha='left')
            
    return fig
    
def plot_graph_with_table(cell_text, row_text, columns):
    plt.style.use('ggplot') 
    fig = plt.figure() 

    # --- Include table --- #
    ax2 = fig.add_subplot(111)
    ax2.axis('off')

    the_table = ax2.table(cellText=cell_text, 
                        rowLabels=row_text, 
                        colLabels=columns, 
                        loc='center right')


## Stage 6 end


